import requests
import urllib.request
from bs4 import BeautifulSoup
import re
import pandas as pd
from pandas import DataFrame
from openpyxl import load_workbook
from numpy import nan
from sys import exit
import time
##portfolio_Analysis
filename = 'portfolio_Analysis.xlsx'

df = pd.read_excel(filename, 'Input')
link = []
link = list (df['Stock link'])
name = list (df['Stock name'])
df1 = pd.read_excel(filename, 'Input Columns')
DataNeeded = df1.to_numpy()


colnum = 0

df3 = pd.read_excel(filename, 'New Output')
datacolumns= df3.to_numpy()
Data_names = datacolumns[0]

CleanDataNames = [x for x in Data_names if str(x) != 'nan']

#DateList = ['Sep2018','Dec2018', 'Mar2019', 'Jun2019', 'Sep2019', 'Dec2019', 'Mar2020', 'Jun2020', 'Sep2020', 'Dec2020', 'Mar2021', 'Jun2021', 'Sep2021']#flip this
DateList = DataNeeded[0]
DateList = DateList.tolist()
DateList = [x for x in DateList if str(x) != 'nan']
#print(DateList)

wb = load_workbook(filename)
ws=wb['New Output']

for i in range(0,len(link)):
    url = link[i]
    print ("URL: "+ url)
    x = 0
    while x == 0:
        try:
            response = requests.get(url)
            x=1
        except:
            wb.save(filename)
            print("Data Saved")
            print("Connection lost trying to reconnect")
            time.sleep(1)
            
    print (response)
    soup = BeautifulSoup(response.text, "html.parser")
    #rownum = 4
    
    for col in range(2,len(DataNeeded)):
        section = (DataNeeded[col,0])
        section = section.lower()
        
        if section == "quaterly results":
            section = "quarters"
        elif section == "profit & loss":
            section = "profit-loss"
        elif section == "balance sheet":
            section = "balance-sheet"
        elif section == "cash flow":
            section = "cash-flows"
        elif section == "shareholding pattern":
            section = "shareholding"
        #print(section)
            
        if section == "ratios":
            htmldata = soup.find('div', {'class':'company-ratios'})
            cleanr = re.compile('<.*?>')
            htmldata = str(htmldata)
            cleanr = re.compile('<.*?>')
            allData = re.sub(cleanr, '', htmldata)
            allData.replace(u'\xa0', '').encode('utf-8')
            listOfData = allData.split('\n')
            
            listOfData = [el.replace('\xa0+','') for el in listOfData]
            cleanlist = []
            for e in listOfData:
                j = e.replace(' ','')
                cleanlist.append(j)
            cleanlist = list(filter(('').__ne__, cleanlist))
            #print("CleanList: ", cleanlist)
        else:
            htmldata = soup.find('section', {'id':section})
            DateB = str(htmldata.find("thead"))
            cleanr = re.compile('<.*?>')
            DateB = re.sub(cleanr, '', DateB)
            DateB = DateB.split('\n')
            DateBList = []
            for date in DateB:
              j = date.replace(' ','')
              DateBList.append(j)
            DateBList = list(filter(('').__ne__, DateBList))

            htmldata = str(htmldata)
            #print(htmldata)
            htmldata = htmldata.replace("""<td class="">\n</td>""", """<td class="">\n          0\n        </td>""")
            #print(htmldata)

            cleanr = re.compile('<.*?>')
            allData = re.sub(cleanr, '', htmldata)
            #print(allData)
            allData.replace(u'\xa0', '').encode('utf-8')
            listOfData = allData.split('\n')
            
            listOfData = [el.replace('\xa0+','') for el in listOfData]
            cleanlist = []
            for e in listOfData:
              j = e.replace(' ','')
              cleanlist.append(j)
            cleanlist = list(filter(('').__ne__, cleanlist))
            #print(cleanlist)
            

        
        for row in range(1,len(DataNeeded[0])):
            data = (DataNeeded[col,row])
            if type(data) != float:
                colu = CleanDataNames.index(data)+1
                #print ("Section: " + section)
                data = data.replace(" ", "")
                #print ("data: "+ data)
                #wb = load_workbook(filename)
                #ws=wb['New Output']
                if section == "ratios":
                    if data in cleanlist:
                        #print('Data in cleanlist')
                        index = cleanlist.index(data)
                        if data == 'MarketCap':
                            end = 'CurrentPrice'
                        elif data == 'High/Low':
                            end == 'StockP/E'
                        elif data == 'StockP/E':
                            end == 'BookValue'
                        if end in cleanlist:
                            endindex = cleanlist.index(end)
                            if endindex < index:
                                endindex = index+2
                        else:
                            if data == 'MarketCap':
                                endindex = index + 3
                            else:
                                endindex = index + 2
                        #print(cleanlist)
                        FinalData = cleanlist[index+1:endindex]
                        if data == 'MarketCap':
                            finaldata = FinalData[-2]
                        elif ('%' in FinalData) & (FinalData != ['%']):
                            #do try except to make finaldata 0 if no % found
                            finaldata = FinfalData.pop('%')
                        else:
                            finaldata = ''.join(FinalData)
                        #print(finaldata)
                        if finaldata == 'BookValue':
                            finfaldata = ''
                        elif FinalData == '%':
                            print ("data: "+ data)
                            print("no Data")
                            finfalData = ['']
                        else:
                            if finaldata.find(',') != -1:
                                finaldata = finaldata.replace(',', '')
                            try:
                                finfaldata = float(finaldata)
                            except:
                                finfaldata = ""

                        for g in range(0,len(DateList)):
                            ws.cell(colnum+g+3, column = colu).value=finfaldata
                        
                    else:
                        print('Data not found')
                        continue
                        
                else:
                    #print("Dates: ", DateList)
                    #print("Dates on website: ", DateBList)
                    matches = list(set(DateBList) & set(DateList))
                    #print("Matches: ",matches)
                    listofmatches = []
                    for item in matches:
                        index_pos_list = DateBList.index(item)
                        listofmatches.append(index_pos_list)
                    listofmatches.sort()
                    #print("list of url date matches: ",listofmatches)
                    listofmatchesURL=[]
                    for item in matches:
                        index_pos_list = DateList.index(item)
                        listofmatchesURL.append(index_pos_list)
                    listofmatchesURL.sort()
                    #print ("list of date matches: ",listofmatchesURL)
                    if data not in cleanlist and data =='Sales':
                        data = 'Revenue'
                    if data not in cleanlist and data == 'OperatingProfit':
                        data = 'FinancingProfit'
                    if data not in cleanlist and data == 'OPM%':
                        data = 'FinancingMargin%'
                    if data in cleanlist:
                        index = cleanlist.index(data)
                        l = len(DateBList) + 1
                        endindex = l+index
                        if data == 'EPSinRs':
                            end = 'RawPDF'
                            endindex = cleanlist.index(end)
                        
                        if 'GrossNPA%' in cleanlist:
                            endindex = cleanlist.index('GrossNPA%')
                        

                        FinalData = cleanlist[index:endindex]
                        
                        
                        if len(FinalData) > l:
                            FinalData = FinalData[0:l]
                        if '%' in FinalData[0]:
                            FinalData = [Data.replace("%", "") for Data in FinalData]
                        FinalData = [Data.replace(",", "") for Data in FinalData]
                        #print(FinalData)
                        remove = []
                        for k in range(1,(len(FinalData))):
                            try:
                                x = float(FinalData[k])
                            except:
                                remove.append(k)
                                
                        #print(remove)
                        for r in remove:
                            #print(r)
                            FinalData.remove(FinalData[r])
                            #FinalData.append("Fix manually")
                        #print(FinalData)

                        #[float(Data) for Data in FinalData]
                        #print("Final Data: ",FinalData)
                        #print("length of datelist:",len(DateList))
                        datecol = CleanDataNames.index("Date")+1
                        namecol = CleanDataNames.index("Company Name")+1
                        
                        for date in range(0,len(DateList)):
                            ws.cell(row=colnum+date+3, column = datecol).value=DateList[date]
                            ws.cell(row=colnum+date+3, column = namecol).value=name[i]
                        for n in listofmatches:
                            u = listofmatches.index(n)
                            x = listofmatchesURL[u]
                            q = colnum+x+3
                            #print(n,u,x,"row: ",q,"column: ",colu)
                            if len(FinalData) >= n+2:
                                if FinalData[n+1] != 'EPSinRs':
                                    data = FinalData[(n+1)]
                                    if data.find(',') != -1:
                                        data = data.replace(',', '')
                                    data = float(data)
                                    #print(data)
                                    ws.cell(row=q,column=colu).value=data
                                else:
                                    #print("Here")
                                    ws.cell(row=q,column=colu).value="Enter Manually"
     
                    else:
                        print ("data: "+ data)
                        print('Data not found')
                        FinalData = []
        #wb.save(filename)
        #print("Data Saved")

    wb.save(filename)
    print("Data Saved")
    colnum = colnum + len(DateList)
